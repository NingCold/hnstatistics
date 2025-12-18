from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from hnstatistics.core.errors import ExportIOError, ExportPathError, ProjectEmptyError
from hnstatistics.core.statistics.model import StatisticsModel


def export_excel(stats: StatisticsModel, file_path: str):
    if not stats.frequency:
        raise ProjectEmptyError()
    
    if not file_path or not isinstance(file_path, str):
        raise ExportPathError(file_path, "Export file path is invalid.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Statistics"

    headers = ["Item", "Frequency", "Probability"]
    ws.append(headers)
    
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for key in stats.frequency:
        freq = stats.frequency[key]
        prob = round(stats.probability[key], 4)
        ws.append([key, freq, prob])
    
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    try:
        wb.save(file_path)
    except OSError as e:
        raise ExportIOError(file_path, str(e))